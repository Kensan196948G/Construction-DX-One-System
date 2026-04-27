import uuid

import pytest
from openpyxl import load_workbook

from apps.audits.models import Audit
from apps.audits.report_generator import (
    generate_audit_excel,
    generate_audit_pdf,
    generate_compliance_summary_excel,
    generate_risk_register_excel,
)
from apps.compliance.models import Control
from apps.risks.models import Risk


def _consume(streaming_response):
    return b"".join(streaming_response.streaming_content)

# ---------------------------------------------------------------------------
# Audit Excel Report
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_audit_excel_report_structure(client, audit_data):
    create_resp = client.post("/api/v1/audits/", data=audit_data, content_type="application/json")
    audit_id = create_resp.json()["id"]
    audit = Audit.objects.get(pk=audit_id)

    for severity, status in [
        ("critical", "open"),
        ("high", "in_remediation"),
        ("medium", "resolved"),
        ("low", "accepted"),
        ("info", "open"),
    ]:
        audit.findings.create(
            title=f"Finding {severity}",
            severity=severity,
            status=status,
            description=f"Description for {severity}",
            recommendation=f"Fix for {severity}",
        )

    findings = audit.findings.all()
    excel_file = generate_audit_excel(audit, findings)
    wb = load_workbook(excel_file)

    sheet_names = wb.sheetnames
    assert "Audit Summary" in sheet_names
    assert "Findings Detail" in sheet_names
    assert "Severity Distribution" in sheet_names

    ws1 = wb["Audit Summary"]
    assert ws1.cell(row=1, column=1).value == "Audit Report"
    assert ws1.cell(row=4, column=1).value == "Title"
    assert ws1.cell(row=4, column=2).value == audit.title

    ws2 = wb["Findings Detail"]
    assert ws2.cell(row=1, column=2).value == "Title"
    assert ws2.cell(row=1, column=3).value == "Severity"

    assert ws2.cell(row=2, column=2).value == "Finding critical"
    assert ws2.cell(row=6, column=2).value == "Finding medium"

    ws3 = wb["Severity Distribution"]
    assert ws3.cell(row=1, column=1).value == "Severity"
    assert ws3.cell(row=1, column=2).value == "Count"
    assert ws3.cell(row=2, column=2).value == 1
    assert ws3.cell(row=6, column=2).value == 1

    assert len(ws3._charts) >= 1


# ---------------------------------------------------------------------------
# Audit PDF Report
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_audit_pdf_report_generation(client, audit_data):
    create_resp = client.post("/api/v1/audits/", data=audit_data, content_type="application/json")
    audit_id = create_resp.json()["id"]
    audit = Audit.objects.get(pk=audit_id)

    for severity in ["critical", "high", "medium", "low", "info"]:
        audit.findings.create(
            title=f"Finding {severity}",
            severity=severity,
            status="open",
            description=f"Description for {severity}",
            recommendation=f"Fix for {severity}",
        )

    findings = audit.findings.all()
    pdf_file = generate_audit_pdf(audit, findings)
    pdf_data = pdf_file.read()

    assert pdf_data.startswith(b"%PDF")
    assert len(pdf_data) > 1000


# ---------------------------------------------------------------------------
# Audit Report HTTP endpoints
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_audit_report_excel_download(client, audit_data):
    create_resp = client.post("/api/v1/audits/", data=audit_data, content_type="application/json")
    audit_id = create_resp.json()["id"]

    response = client.get(f"/api/v1/audits/{audit_id}/report/excel/")
    assert response.status_code == 200
    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response["Content-Disposition"].startswith("attachment; filename=")
    assert response["Content-Disposition"].endswith(".xlsx\"")
    assert len(_consume(response)) > 100


@pytest.mark.django_db
def test_audit_report_pdf_download(client, audit_data):
    create_resp = client.post("/api/v1/audits/", data=audit_data, content_type="application/json")
    audit_id = create_resp.json()["id"]

    response = client.get(f"/api/v1/audits/{audit_id}/report/pdf/")
    assert response.status_code == 200
    assert response["Content-Type"] == "application/pdf"
    assert response["Content-Disposition"].startswith("attachment; filename=")
    assert response["Content-Disposition"].endswith(".pdf\"")
    assert _consume(response).startswith(b"%PDF")


@pytest.mark.django_db
def test_audit_report_not_found(client):
    fake_id = uuid.uuid4()
    response = client.get(f"/api/v1/audits/{fake_id}/report/excel/")
    assert response.status_code == 404

    response = client.get(f"/api/v1/audits/{fake_id}/report/pdf/")
    assert response.status_code == 404


# ---------------------------------------------------------------------------
# Compliance Excel Report
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_compliance_excel_report(client, control_data):
    client.post("/api/v1/compliance/", data=control_data, content_type="application/json")

    controls = Control.objects.all()
    excel_file = generate_compliance_summary_excel(controls)
    wb = load_workbook(excel_file)

    assert "Compliance Summary" in wb.sheetnames
    ws = wb["Compliance Summary"]
    assert ws.cell(row=1, column=1).value == "Compliance Summary Report"
    assert ws.cell(row=4, column=1).value == "Organizational Controls"


@pytest.mark.django_db
def test_compliance_excel_report_http(client, control_data):
    client.post("/api/v1/compliance/", data=control_data, content_type="application/json")

    response = client.get("/api/v1/compliance/report/excel/")
    assert response.status_code == 200
    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert len(_consume(response)) > 100


# ---------------------------------------------------------------------------
# Risk Excel Report
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_risk_excel_report(client, risk_data):
    client.post("/api/v1/risks/", data=risk_data, content_type="application/json")

    risks = Risk.objects.all()
    excel_file = generate_risk_register_excel(risks)
    wb = load_workbook(excel_file)

    assert "Risk Register" in wb.sheetnames
    ws = wb["Risk Register"]
    assert ws.cell(row=1, column=1).value == "Risk Register Report"
    assert ws.cell(row=4, column=2).value == "Title"


@pytest.mark.django_db
def test_risk_excel_report_http(client, risk_data):
    client.post("/api/v1/risks/", data=risk_data, content_type="application/json")

    response = client.get("/api/v1/risks/report/excel/")
    assert response.status_code == 200
    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert len(_consume(response)) > 100


# ---------------------------------------------------------------------------
# Edge Cases
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_audit_excel_no_findings(client, audit_data):
    create_resp = client.post("/api/v1/audits/", data=audit_data, content_type="application/json")
    audit_id = create_resp.json()["id"]
    audit = Audit.objects.get(pk=audit_id)

    findings = audit.findings.all()
    excel_file = generate_audit_excel(audit, findings)
    wb = load_workbook(excel_file)

    ws2 = wb["Findings Detail"]
    assert ws2.cell(row=2, column=1).value is None

    ws3 = wb["Severity Distribution"]
    for row in range(2, 7):
        assert ws3.cell(row=row, column=2).value == 0


@pytest.mark.django_db
def test_audit_pdf_no_findings(client, audit_data):
    create_resp = client.post("/api/v1/audits/", data=audit_data, content_type="application/json")
    audit_id = create_resp.json()["id"]
    audit = Audit.objects.get(pk=audit_id)

    pdf_file = generate_audit_pdf(audit, audit.findings.all())
    assert pdf_file.read().startswith(b"%PDF")


@pytest.mark.django_db
def test_risk_excel_heatmap_coloring(client, risk_data):
    for likelihood, impact in [(5, 5), (4, 3), (3, 2), (2, 1), (1, 1)]:
        data = {**risk_data, "title": f"Risk L{likelihood}I{impact}", "likelihood": likelihood, "impact": impact}
        client.post("/api/v1/risks/", data=data, content_type="application/json")

    risks = Risk.objects.all()
    excel_file = generate_risk_register_excel(risks)
    wb = load_workbook(excel_file)
    ws = wb["Risk Register"]

    risk_levels = []
    for row in range(5, 5 + 5):
        risk_levels.append(ws.cell(row=row, column=7).value)

    assert "Critical" in risk_levels
    assert "Low" in risk_levels
