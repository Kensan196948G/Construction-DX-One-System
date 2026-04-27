from __future__ import annotations

from datetime import datetime
from io import BytesIO

from django.db.models import QuerySet
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.audits.models import Audit, Finding
from apps.compliance.models import Control
from apps.risks.models import Risk

SEVERITY_COLORS = {
    "critical": "FF0000",
    "high": "FF6600",
    "medium": "FFD700",
    "low": "92D050",
    "info": "4472C4",
}

SEVERITY_ORDER = ["critical", "high", "medium", "low", "info"]

STATUS_COLORS = {
    "open": "FF0000",
    "in_remediation": "FF6600",
    "resolved": "92D050",
    "accepted": "4472C4",
}

RISK_LEVEL_COLORS = {
    "critical": "FF0000",
    "high": "FF6600",
    "medium": "FFD700",
    "low": "92D050",
}

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_header_style(ws, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _apply_body_style(ws, start_row: int, end_row: int, num_cols: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER


def _set_column_widths(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_audit_excel(audit: Audit, findings: QuerySet | list[Finding]) -> BytesIO:
    wb = Workbook()

    # --- Sheet 1: Audit Summary ---
    ws1 = wb.active
    ws1.title = "Audit Summary"
    ws1.sheet_properties.tabColor = "2F5496"

    ws1.cell(row=1, column=1, value="Audit Report").font = TITLE_FONT
    ws1.merge_cells("A1:D1")

    summary_fields = [
        ("Title", audit.title),
        ("Scope", audit.scope),
        ("Auditor", audit.auditor),
        ("Status", audit.get_status_display() if audit.status else ""),
        ("Planned Date", str(audit.planned_date or "")),
        ("Completed Date", str(audit.completed_date or "")),
        ("Total Findings", str(findings.count())),
    ]

    ws1.cell(row=3, column=1, value="Field").font = HEADER_FONT
    ws1.cell(row=3, column=1).fill = HEADER_FILL
    ws1.cell(row=3, column=1).border = THIN_BORDER
    ws1.cell(row=3, column=2, value="Value").font = HEADER_FONT
    ws1.cell(row=3, column=2).fill = HEADER_FILL
    ws1.cell(row=3, column=2).border = THIN_BORDER

    for i, (field, value) in enumerate(summary_fields):
        row = 4 + i
        ws1.cell(row=row, column=1, value=field).font = Font(name="Calibri", bold=True, size=10)
        ws1.cell(row=row, column=1).border = THIN_BORDER
        ws1.cell(row=row, column=2, value=value).font = BODY_FONT
        ws1.cell(row=row, column=2).border = THIN_BORDER

    _set_column_widths(ws1, [20, 60])

    # --- Sheet 2: Findings Detail ---
    ws2 = wb.create_sheet("Findings Detail")
    ws2.sheet_properties.tabColor = "C00000"

    headers = ["#", "Title", "Severity", "Status", "Description", "Recommendation"]
    for col, header in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=header)
    _apply_header_style(ws2, 1, len(headers))

    for i, finding in enumerate(findings):
        row = 2 + i
        ws2.cell(row=row, column=1, value=i + 1)
        ws2.cell(row=row, column=2, value=finding.title)
        ws2.cell(row=row, column=3, value=finding.get_severity_display() if finding.severity else finding.severity)
        ws2.cell(row=row, column=4, value=finding.get_status_display() if finding.status else finding.status)
        ws2.cell(row=row, column=5, value=finding.description)
        ws2.cell(row=row, column=6, value=finding.recommendation)

        sev_color = SEVERITY_COLORS.get(finding.severity, "000000")
        ws2.cell(row=row, column=3).fill = PatternFill(
            start_color=sev_color, end_color=sev_color, fill_type="solid"
        )
        ws2.cell(row=row, column=3).font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    _apply_body_style(ws2, 2, 1 + len(findings), len(headers))
    _set_column_widths(ws2, [5, 40, 12, 18, 50, 50])

    # --- Sheet 3: Severity Distribution ---
    ws3 = wb.create_sheet("Severity Distribution")
    ws3.sheet_properties.tabColor = "00B050"

    severity_counts: dict[str, int] = {}
    for sev in SEVERITY_ORDER:
        severity_counts[sev] = 0
    for finding in findings:
        severity_counts[finding.severity] = severity_counts.get(finding.severity, 0) + 1

    headers3 = ["Severity", "Count"]
    for col, header in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=header)
    _apply_header_style(ws3, 1, len(headers3))

    row = 2
    for sev in SEVERITY_ORDER:
        count = severity_counts.get(sev, 0)
        ws3.cell(row=row, column=1, value=sev.capitalize())
        ws3.cell(row=row, column=2, value=count)
        sev_color = SEVERITY_COLORS.get(sev, "000000")
        ws3.cell(row=row, column=1).fill = PatternFill(
            start_color=sev_color, end_color=sev_color, fill_type="solid"
        )
        ws3.cell(row=row, column=1).font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        row += 1

    _apply_body_style(ws3, 2, row - 1, len(headers3))

    chart = BarChart()
    chart.type = "col"
    chart.title = "Findings by Severity"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Severity"
    chart.style = 10

    data = Reference(ws3, min_col=2, min_row=1, max_row=row - 1)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 16
    chart.height = 10

    ws3.add_chart(chart, "D1")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _severity_to_risk_level(severity: str) -> str:
    if severity == "critical":
        return "critical"
    if severity == "high":
        return "high"
    if severity == "medium":
        return "medium"
    return "low"


def _status_summary(findings: list[Finding]) -> dict[str, int]:
    summary: dict[str, int] = {}
    for finding in findings:
        summary[finding.status] = summary.get(finding.status, 0) + 1
    return summary


AUDIT_PDF_TEMPLATE = """<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<style>
  @page {{
    size: A4;
    margin: 2cm;
    @top-center {{
      content: element(pageHeader);
    }}
  }}
  body {{
    font-family: 'Noto Sans JP', 'Hiragino Kaku Gothic ProN', 'MS Gothic', sans-serif;
    font-size: 10pt;
    color: #333;
    line-height: 1.6;
  }}
  .cover-page {{
    page-break-after: always;
    text-align: center;
    padding-top: 120px;
  }}
  .cover-page h1 {{
    font-size: 24pt;
    color: #2F5496;
    margin-bottom: 8px;
  }}
  .cover-page .subtitle {{
    font-size: 14pt;
    color: #666;
    margin-bottom: 40px;
  }}
  .cover-page .meta {{
    font-size: 11pt;
    color: #555;
    margin-top: 60px;
  }}
  .cover-page .meta table {{
    margin: 0 auto;
    border-collapse: collapse;
  }}
  .cover-page .meta td {{
    padding: 6px 16px;
    text-align: left;
  }}
  .cover-page .meta td.label {{
    font-weight: bold;
    text-align: right;
    color: #2F5496;
  }}
  .section-title {{
    font-size: 16pt;
    color: #2F5496;
    border-bottom: 2px solid #2F5496;
    padding-bottom: 4px;
    margin-top: 30px;
    margin-bottom: 16px;
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 20px;
  }}
  th, td {{
    border: 1px solid #ccc;
    padding: 6px 10px;
    text-align: left;
    vertical-align: top;
  }}
  th {{
    background-color: #2F5496;
    color: white;
    font-weight: bold;
    text-align: center;
  }}
  tr:nth-child(even) {{
    background-color: #f8f9fa;
  }}
  .severity-critical {{ background-color: #FFCCCC; }}
  .severity-high {{ background-color: #FFE0B2; }}
  .severity-medium {{ background-color: #FFF9C4; }}
  .severity-low {{ background-color: #C8E6C9; }}
  .severity-info {{ background-color: #BBDEFB; }}
  .badge {{
    display: inline-block;
    padding: 2px 10px;
    border-radius: 10px;
    color: white;
    font-weight: bold;
    font-size: 9pt;
  }}
  .badge-critical {{ background-color: #FF0000; }}
  .badge-high {{ background-color: #FF6600; }}
  .badge-medium {{ background-color: #DAA520; }}
  .badge-low {{ background-color: #4CAF50; }}
  .badge-info {{ background-color: #4472C4; }}
  .summary-box {{
    display: flex;
    justify-content: space-around;
    margin: 20px 0;
    flex-wrap: wrap;
  }}
  .summary-item {{
    text-align: center;
    padding: 12px 20px;
    background: #f0f4ff;
    border-radius: 8px;
    min-width: 100px;
  }}
  .summary-item .number {{
    font-size: 24pt;
    font-weight: bold;
    color: #2F5496;
  }}
  .summary-item .label {{
    font-size: 9pt;
    color: #666;
  }}
  .footer {{
    text-align: center;
    font-size: 8pt;
    color: #999;
    margin-top: 40px;
    border-top: 1px solid #eee;
    padding-top: 10px;
  }}
</style>
</head>
<body>

<div class="cover-page">
  <h1>Audit Report</h1>
  <div class="subtitle">{audit_title}</div>
  <div class="meta">
    <table>
      <tr><td class="label">Status</td><td>{audit_status}</td></tr>
      <tr><td class="label">Auditor</td><td>{auditor}</td></tr>
      <tr><td class="label">Scope</td><td>{scope}</td></tr>
      <tr><td class="label">Planned Date</td><td>{planned_date}</td></tr>
      <tr><td class="label">Completed Date</td><td>{completed_date}</td></tr>
      <tr><td class="label">Total Findings</td><td>{total_findings}</td></tr>
      <tr><td class="label">Generated</td><td>{generated_at}</td></tr>
    </table>
  </div>
</div>

<h2 class="section-title">Status Summary</h2>
<div class="summary-box">
  <div class="summary-item">
    <div class="number">{total_findings}</div>
    <div class="label">Total Findings</div>
  </div>
  {status_boxes}
</div>

<h2 class="section-title">Findings Detail</h2>
<table>
  <thead>
    <tr>
      <th style="width:5%">#</th>
      <th style="width:30%">Title</th>
      <th style="width:10%">Severity</th>
      <th style="width:14%">Status</th>
      <th style="width:41%">Description / Recommendation</th>
    </tr>
  </thead>
  <tbody>
    {finding_rows}
  </tbody>
</table>

<div class="footer">
  Generated by CGRC System | {generated_at}
</div>

</body>
</html>
"""


def generate_audit_pdf(audit: Audit, findings: QuerySet | list[Finding]) -> BytesIO:
    from weasyprint import HTML

    findings_list = list(findings)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    status_summary = _status_summary(findings_list)
    status_labels = {
        "open": "Open",
        "in_remediation": "In Remediation",
        "resolved": "Resolved",
        "accepted": "Accepted",
    }
    status_boxes = ""
    for status_key, status_label in status_labels.items():
        count = status_summary.get(status_key, 0)
        status_boxes += (
            f'<div class="summary-item">'
            f'<div class="number">{count}</div>'
            f'<div class="label">{status_label}</div>'
            f"</div>\n"
        )

    finding_rows = ""
    for i, finding in enumerate(findings_list):
        sev_lower = finding.severity or "info"
        sev_display = finding.get_severity_display() if finding.severity else ""
        status_display = finding.get_status_display() if finding.status else ""
        finding_rows += (
            f'<tr class="severity-{sev_lower}">'
            f"<td>{i + 1}</td>"
            f"<td>{_escape_html(finding.title)}</td>"
            f'<td><span class="badge badge-{sev_lower}">{_escape_html(sev_display)}</span></td>'
            f"<td>{_escape_html(status_display)}</td>"
            f"<td><b>Description:</b> {_escape_html(finding.description)}<br>"
            f"<b>Recommendation:</b> {_escape_html(finding.recommendation)}</td>"
            f"</tr>\n"
        )

    status_display = audit.get_status_display() if audit.status else ""

    html = AUDIT_PDF_TEMPLATE.format(
        audit_title=_escape_html(audit.title),
        audit_status=_escape_html(status_display),
        auditor=_escape_html(audit.auditor),
        scope=_escape_html(audit.scope),
        planned_date=str(audit.planned_date or ""),
        completed_date=str(audit.completed_date or ""),
        total_findings=str(len(findings_list)),
        generated_at=now_str,
        status_boxes=status_boxes,
        finding_rows=finding_rows,
    )

    output = BytesIO()
    HTML(string=html).write_pdf(output)
    output.seek(0)
    return output


def _escape_html(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#x27;")
    )


def generate_compliance_summary_excel(controls: QuerySet | list[Control]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Summary"
    ws.sheet_properties.tabColor = "2F5496"

    ws.cell(row=1, column=1, value="Compliance Summary Report").font = TITLE_FONT
    ws.merge_cells("A1:E1")

    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = BODY_FONT
    ws.merge_cells("A2:E2")

    domain_map: dict[str, list[Control]] = {}
    for c in controls:
        domain = c.domain or "Other"
        domain_map.setdefault(domain, []).append(c)

    row = 4
    for domain_name, domain_controls in domain_map.items():
        start_row = row
        ws.cell(row=row, column=1, value=domain_name).font = Font(
            name="Calibri", bold=True, size=12, color="2F5496"
        )
        ws.merge_cells(f"A{row}:E{row}")
        row += 1

        headers = [
            "Control ID",
            "Title",
            "Applicability",
            "Implementation Status",
            "Justification",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        _apply_header_style(ws, row, len(headers))
        row += 1

        for ctrl in domain_controls:
            ws.cell(row=row, column=1, value=ctrl.control_number)
            ws.cell(row=row, column=2, value=ctrl.title)
            ws.cell(row=row, column=3, value=ctrl.get_applicability_display() if ctrl.applicability else "")
            ws.cell(row=row, column=4, value=ctrl.get_implementation_status_display() if ctrl.implementation_status else "")

            impl_color = _implementation_status_color(ctrl.implementation_status)
            if impl_color:
                ws.cell(row=row, column=4).fill = PatternFill(
                    start_color=impl_color, end_color=impl_color, fill_type="solid"
                )
                ws.cell(row=row, column=4).font = Font(
                    name="Calibri", bold=True, color="FFFFFF", size=10
                )

            ws.cell(row=row, column=5, value=ctrl.justification)
            row += 1

        _apply_body_style(ws, start_row + 1, row - 1, len(headers))
        row += 1  # blank row between domains

    _set_column_widths(ws, [14, 40, 16, 22, 50])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _implementation_status_color(status: str) -> str | None:
    colors = {
        "verified": "00B050",
        "implemented": "92D050",
        "in_progress": "FFD700",
        "not_started": "FF6600",
        "not_applicable": "A0A0A0",
    }
    return colors.get(status)


def _risk_level(likelihood: int, impact: int) -> str:
    score = likelihood * impact
    if score >= 20:
        return "critical"
    if score >= 12:
        return "high"
    if score >= 6:
        return "medium"
    return "low"


def generate_risk_register_excel(risks: QuerySet | list[Risk]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Register"
    ws.sheet_properties.tabColor = "C00000"

    ws.cell(row=1, column=1, value="Risk Register Report").font = TITLE_FONT
    ws.merge_cells("A1:H1")

    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = BODY_FONT
    ws.merge_cells("A2:H2")

    headers = [
        "#",
        "Title",
        "Category",
        "Likelihood",
        "Impact",
        "Risk Score",
        "Risk Level",
        "Status",
    ]
    row = 4
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    _apply_header_style(ws, row, len(headers))
    row += 1

    for i, risk in enumerate(risks):
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=risk.title)
        ws.cell(row=row, column=3, value=risk.category)
        ws.cell(row=row, column=4, value=risk.likelihood)
        ws.cell(row=row, column=5, value=risk.impact)
        ws.cell(row=row, column=6, value=risk.risk_score)

        level = _risk_level(risk.likelihood, risk.impact)
        ws.cell(row=row, column=7, value=level.capitalize())
        ws.cell(row=row, column=8, value=risk.get_status_display() if risk.status else risk.status)

        risk_color = RISK_LEVEL_COLORS.get(level, "FFFFFF")
        ws.cell(row=row, column=7).fill = PatternFill(
            start_color=risk_color, end_color=risk_color, fill_type="solid"
        )
        ws.cell(row=row, column=7).font = Font(
            name="Calibri", bold=True, color="FFFFFF", size=10
        )

        row += 1

    _apply_body_style(ws, 5, row - 1, len(headers))
    _set_column_widths(ws, [5, 40, 18, 12, 10, 12, 14, 16])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
